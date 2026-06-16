#!/usr/bin/env python3
"""
gen_cv_lists.py  —  Update Colonius.tex from the master spreadsheet.

Usage:
    python3 gen_cv_lists.py

Reads:  data/Colonius Group.xlsx
Writes: Colonius.tex

Two things are updated:
  1. List sections (between BEGIN_GENERATED / END_GENERATED markers):
       Awards, Teaching, Committees, Professional Activities, Patents,
       Invited Lectures, Doctoral Students (committee)

  2. \\nocite{} lines in each \\begin{refsection} block:
       Submitted Articles, Journal Articles, Conference Papers,
       Book Chapters, Doctoral students (as advisor)
"""

from __future__ import annotations   # type hints are strings → runs on Python 3.7+

import re
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl not found — run: pip install openpyxl")

import sheet_data

# ── Paths ──────────────────────────────────────────────────────────────────────
HERE = Path(__file__).parent
TEX  = HERE / "cv" / "Colonius.tex"
BIB  = HERE / "cv" / "colonius.bib"   # bibliography consumed by Colonius.tex

# ══════════════════════════════════════════════════════════════════════════════
# PART 0 — LaTeX processing layer (shared by CV lists and the bibliography)
#
# Contract: the spreadsheet holds PLAIN TEXT and is treated as IMMUTABLE input.
# This layer does ALL LaTeX conversion (escape specials, Unicode → TeX,
# capitalisation protection).  If a cell cannot be converted cleanly it raises
# an error pinpointing the cell — the build never patches the input.
# ══════════════════════════════════════════════════════════════════════════════

# Unicode → LaTeX, applied to every typeset field.  Covers accented letters,
# dashes, non-breaking space and smart quotes.  Does NOT include the ASCII
# specials & % # _  — those are handled by escape_specials() for text fields and
# left untouched in verbatim URL/DOI fields.
UNICODE_MAP = {
    "\xa0": " ",
    "–": "--", "—": "---", "‐": "-", "‒": "-",
    "‘": "`", "’": "'", "“": "``", "”": "''", '"': "''",
    "ä": r'{\"a}', "ë": r'{\"e}', "ï": r'{\"i}', "ö": r'{\"o}', "ü": r'{\"u}', "ÿ": r'{\"y}',
    "Ä": r'{\"A}', "Ë": r'{\"E}', "Ï": r'{\"I}', "Ö": r'{\"O}', "Ü": r'{\"U}',
    "á": r"{\'a}", "é": r"{\'e}", "í": r"{\'i}", "ó": r"{\'o}", "ú": r"{\'u}", "ý": r"{\'y}",
    "Á": r"{\'A}", "É": r"{\'E}", "Í": r"{\'I}", "Ó": r"{\'O}", "Ú": r"{\'U}", "Ý": r"{\'Y}",
    "ń": r"{\'n}", "Ń": r"{\'N}",
    "à": r"{\`a}", "è": r"{\`e}", "ì": r"{\`i}", "ò": r"{\`o}", "ù": r"{\`u}",
    "À": r"{\`A}", "È": r"{\`E}", "Ì": r"{\`I}", "Ò": r"{\`O}", "Ù": r"{\`U}",
    "â": r"{\^a}", "ê": r"{\^e}", "î": r"{\^i}", "ô": r"{\^o}", "û": r"{\^u}",
    "Â": r"{\^A}", "Ê": r"{\^E}", "Î": r"{\^I}", "Ô": r"{\^O}", "Û": r"{\^U}",
    "ã": r"{\~a}", "ñ": r"{\~n}", "õ": r"{\~o}", "Ã": r"{\~A}", "Ñ": r"{\~N}", "Õ": r"{\~O}",
    "ç": r"{\c{c}}", "Ç": r"{\c{C}}",
    "ø": r"{\o}", "Ø": r"{\O}", "å": r"{\aa}", "Å": r"{\AA}",
    "æ": r"{\ae}", "Æ": r"{\AE}", "œ": r"{\oe}", "Œ": r"{\OE}", "ß": r"{\ss}",
    "š": r"{\v{s}}", "Š": r"{\v{S}}", "č": r"{\v{c}}", "Č": r"{\v{C}}",
    "ž": r"{\v{z}}", "Ž": r"{\v{Z}}", "ř": r"{\v{r}}", "Ř": r"{\v{R}}",
}

# Hand-typed LaTeX has no place in plain-text input.
_FORBIDDEN = set("\\{}$")

def _fail(where: str, msg: str, value: str):
    sys.exit(f"\nINPUT ERROR  {where}\n  {msg}\n  Cell value: {value!r}\n"
             f"  → Fix the spreadsheet. Input is immutable; the build will not patch it.")

def _validate(text: str, where: str):
    """Raise if text contains hand-typed LaTeX or an unmapped non-ASCII char."""
    for ch in text:
        if ch in _FORBIDDEN:
            _fail(where, f"contains forbidden LaTeX character {ch!r}", text)
        if ord(ch) > 127 and ch not in UNICODE_MAP:
            _fail(where, f"contains unmapped non-ASCII character {ch!r} "
                         f"(U+{ord(ch):04X}); add it to UNICODE_MAP or fix the cell", text)

def _apply_unicode(text: str) -> str:
    for k, v in UNICODE_MAP.items():
        text = text.replace(k, v)
    return text

def escape_specials(text: str) -> str:
    """Escape ASCII LaTeX specials for text fields (not URLs/DOIs)."""
    return (text.replace("&", r"\&").replace("%", r"\%")
                .replace("#", r"\#").replace("_", r"\_"))

def tex_text(value: str, where: str) -> str:
    """Plain-text field → LaTeX: validate, escape specials, convert Unicode."""
    _validate(value, where)
    return _apply_unicode(escape_specials(value))

def tex_verbatim(value: str, where: str) -> str:
    """URL/DOI field → LaTeX: keep _ and # raw, but escape the bib-breaking & %."""
    _validate(value, where)
    return _apply_unicode(value.replace("&", r"\&").replace("%", r"\%"))

def protect_caps(title: str) -> str:
    """Brace capitalised words so biblatex cannot downcase them (titles only).
    Assumes sentence case: the first word of the title and of each ':'-subtitle
    are left unprotected; every later capital is wrapped in {}."""
    tokens = re.split(r"(\s+)", title)
    first_word, after_colon = True, False
    out = []
    for tok in tokens:
        if re.match(r"\s+", tok):
            out.append(tok)
            continue
        sentence_start = first_word or after_colon
        after_colon = tok.rstrip().endswith(":")
        first_word = False
        if sentence_start:
            out.append(tok)
            continue
        out.append(re.sub(r"[A-Za-z]",
                          lambda m: "{" + m.group(0) + "}" if m.group(0).isupper() else m.group(0),
                          tok))
    return "".join(out)

def tex_title(value: str, where: str) -> str:
    """Bibliography title → LaTeX with capitalisation protection, brace-wrapped."""
    return "{" + protect_caps(tex_text(value, where)) + "}"

# ══════════════════════════════════════════════════════════════════════════════
# PART 1 — List sections from CV Data sheet
# ══════════════════════════════════════════════════════════════════════════════

SECTION_KEYS = {
    "Awards":                                   "awards",
    "Teaching-Caltech":                         "teaching-caltech",
    "Teaching - Short Courses":                 "teaching-shortcourses",
    "Caltech Committees":                       "committees",
    "Other Professional Activities":            "activities",
    "Patents":                                  "patents",
    "Invited Lectures":                         "lectures",
    "Doctoral Students (member of committee)":  "committee-students",
}

def fmt_plain(items):
    return [f"\\item {t}" for t in items]

def fmt_activities(items):
    """Apply \\textsc{org}: rest when item contains a colon."""
    out = []
    for t in items:
        if ":" in t:
            org, rest = t.split(":", 1)
            out.append(f"\\item \\textsc{{{org.strip()}}}: {rest.strip()}")
        else:
            out.append(f"\\item {t}")
    return out

FORMATTERS = {
    "awards":               fmt_plain,
    "teaching-caltech":     fmt_plain,
    "teaching-shortcourses": fmt_plain,
    "committees":           fmt_plain,
    "activities":           fmt_activities,
    "patents":              fmt_plain,
    "lectures":             fmt_plain,
    "committee-students":   fmt_plain,
}

def update_list_sections(tex: str, wb) -> str:
    ws = wb["CV Data"]
    sections: dict[str, list[str]] = {}
    current = None
    for rownum, row in enumerate(ws.iter_rows(values_only=True), start=1):
        a = str(row[0]).strip() if row[0] else ""
        b = str(row[1]).strip() if row[1] else ""
        if a and a in SECTION_KEYS:
            current = a
            sections[current] = []
        elif current and b:
            where = f"CV Data!B{rownum} ({current})"
            sections[current].append(tex_text(b, where))

    replaced = 0
    for section_name, key in SECTION_KEYS.items():
        items = sections.get(section_name, [])
        if not items:
            print(f"  WARNING: no items for '{section_name}'")
            continue
        item_lines = "\n".join(FORMATTERS[key](items)) + "\n"
        pattern = (r"(% BEGIN_GENERATED: " + re.escape(key) + r"\n)"
                   r".*?"
                   r"(% END_GENERATED: " + re.escape(key) + r")")
        def _repl(m, nc=item_lines):
            return m.group(1) + nc + m.group(2)
        tex, n = re.subn(pattern, _repl, tex, flags=re.DOTALL)
        if n == 0:
            print(f"  WARNING: marker BEGIN_GENERATED: {key} not found")
        else:
            replaced += 1
            print(f"  ✓ list: {key}  ({len(items)} items)")

    return tex

# ══════════════════════════════════════════════════════════════════════════════
# PART 2 — \\nocite{} lines from Publications sheet
# ══════════════════════════════════════════════════════════════════════════════

# Maps \\printbibliography title → entry type in spreadsheet
NOCITE_SECTIONS = {
    "Submitted Articles":             "misc",
    "Journal Articles":               "article",
    "Conference Papers":              "inproceedings",
    "Book Chapters":                  "incollection",
    "Doctoral students (as advisor)": "phdthesis",
}

# Column-name → spreadsheet column letter, filled by load_publications() so bib
# errors can name the exact cell (e.g. "Publications!E35 (Title)").
PUB_COL: dict[str, str] = {}

def load_publications(wb) -> list[dict]:
    """Read Publications sheet, return list of dicts keyed by header row.
    Each record carries "_row" (its spreadsheet row number) for error messages."""
    ws = wb["Publications"]
    headers = None
    rows = []
    for rownum, raw in enumerate(ws.iter_rows(values_only=True), start=1):
        if raw[0] == "Entry type":
            headers = [str(h).strip() if h else "" for h in raw]
            PUB_COL.clear()
            for i, h in enumerate(headers):
                if h:
                    PUB_COL[h] = get_column_letter(i + 1)
            continue
        if headers is None:
            continue
        rec = {}
        for i, v in enumerate(raw):
            if i < len(headers):
                rec[headers[i]] = str(v).strip() if v and not hasattr(v, "strftime") else ""
        if rec.get("Tag") and rec.get("Entry type"):
            rec["_row"] = rownum
            rows.append(rec)
    return rows

def update_nocites(tex: str, wb, today: str) -> str:
    pubs = load_publications(wb)
    print(f"  Publications loaded: {len(pubs)} entries")

    # Build per-section key lists sorted year-descending, then first-author
    def sort_key(r):
        try:    yr = -int(float(r.get("Year", 0) or 0))
        except: yr = 0
        auth = r.get("Author", "").split("/")[0].split(",")[0].lower()
        return (yr, auth)

    pubs_sorted = sorted(pubs, key=sort_key)

    nocites: dict[str, str] = {}
    for title, etype in NOCITE_SECTIONS.items():
        keys = [r["Tag"] for r in pubs_sorted
                if r.get("Entry type", "").lower() == etype and r.get("Tag")]
        if keys:
            nocites[title] = ",".join(keys)
            print(f"  ✓ nocite: {title}  ({len(keys)} keys)")
        else:
            print(f"  WARNING: no entries for nocite section '{title}'")

    # Replace \\nocite{} lines within each refsection
    lines = tex.splitlines(keepends=True)
    in_refsection = False
    section_title = None
    nocite_idx    = None
    autogen_idx   = None
    replacements  = []   # (start_line, end_line_exclusive, new_text)

    for i, line in enumerate(lines):
        s = line.strip()
        if r"\begin{refsection}" in line:
            in_refsection = True; section_title = None
            nocite_idx = None; autogen_idx = None
        elif r"\end{refsection}" in line:
            if section_title and nocite_idx is not None and section_title in nocites:
                new = f"% AUTO-GENERATED {today}\n\\nocite{{{nocites[section_title]}}}\n"
                start = autogen_idx if autogen_idx is not None else nocite_idx
                replacements.append((start, nocite_idx + 1, new))
            in_refsection = False; section_title = None
            nocite_idx = None; autogen_idx = None
        elif in_refsection:
            m = re.search(r'\\printbibliography\[.*?title=\{([^}]+)\}', line)
            if m:
                section_title = m.group(1)
            if s.startswith("% AUTO-GENERATED"):
                autogen_idx = i
            elif s.startswith(r"\nocite{") and nocite_idx is None:
                # Confirm autogen_idx is immediately before (no real content between)
                if autogen_idx is not None:
                    between = [l.strip() for l in lines[autogen_idx+1:i]
                               if l.strip() and not l.strip().startswith("%")]
                    if between:
                        autogen_idx = None
                nocite_idx = i
            elif s and not s.startswith("%"):
                if nocite_idx is None:
                    autogen_idx = None

    for start, end, new_text in sorted(replacements, reverse=True):
        lines[start:end] = [new_text]

    print(f"  Updated {len(replacements)} \\nocite blocks")
    return "".join(lines)

# ══════════════════════════════════════════════════════════════════════════════
# PART 3 — colonius.bib from the Publications sheet
#
# Built from the SAME load_publications() data that drives the \nocite{} lines,
# so citation keys can never drift between the two.
# ══════════════════════════════════════════════════════════════════════════════

# entry type → [(biblatex field, spreadsheet column, kind)]
#   kind: "text" (escape + Unicode), "verbatim" (URL/DOI), "name" (text, {}-wrapped)
BIBTEX_FIELDS = {
    "article": [
        ("journal",   "Publication Title", "name"),
        ("publisher", "Publisher",         "name"),
        ("volume",    "Volume",            "text"),
        ("number",    "Issue/Number",      "text"),
        ("pages",     "Page Range",        "text"),
        ("url",       "Persistent URL",    "verbatim"),
        ("doi",       "DOI",               "verbatim"),
        ("issn",      "ISSN",              "text"),
    ],
    "inproceedings": [
        ("booktitle", "Publication Title", "name"),
        ("publisher", "Publisher",         "name"),
        ("pages",     "Page Range",        "text"),
        ("url",       "Persistent URL",    "verbatim"),
        ("doi",       "DOI",               "verbatim"),
    ],
    "incollection": [
        ("booktitle", "Publication Title", "name"),
        ("editor",    "Editor",            "name"),
        ("publisher", "Publisher",         "name"),
        ("pages",     "Page Range",        "text"),
    ],
    "misc": [
        ("howpublished", "Publication Title", "name"),
        ("url",          "Persistent URL",    "verbatim"),
        ("note",         "Submitted to",      "text"),
    ],
    "phdthesis": [
        ("school", "School",         "name"),
        ("url",    "Persistent URL", "verbatim"),
    ],
}

def _wrap_field(key: str, value: str) -> str:
    """Wrap long field values so no line exceeds ~250 chars: biber's C parser
    (libbtparse) has a fixed input buffer and segfaults on very long lines."""
    line = f"  {key} = {{{value}}},"
    if len(line) <= 250:
        return line
    if key == "author":
        sep = " and\n" + " " * 15  # align continuation under the opening brace
        wrapped = sep.join(value.split(" and "))
        return f"  {key} = {{{wrapped}}},"
    import textwrap
    indent = " " * (len(key) + 7)
    wrapped = textwrap.fill(value, width=250, subsequent_indent=indent)
    return f"  {key} = {{{wrapped}}},"

def make_bibtex_entry(rec: dict) -> str:
    etype = rec.get("Entry type", "").lower()
    tag   = rec.get("Tag", "")
    if not tag:
        return ""
    row = rec.get("_row", "?")
    def where(col):
        return f"Publications!{PUB_COL.get(col, '?')}{row} ({col})"

    try:
        year = str(int(float(rec.get("Year", ""))))
    except ValueError:
        year = rec.get("Year", "")

    author_raw = rec.get("Author", "")
    _validate(author_raw, where("Author"))
    authors = [a.strip() for a in author_raw.split("/") if a.strip()]
    author_tex = _apply_unicode(escape_specials(" and ".join(authors)))

    fields = [
        ("year",   year),
        ("author", author_tex),
        ("title",  tex_title(rec.get("Title", ""), where("Title"))),
    ]
    for key, col, kind in BIBTEX_FIELDS.get(etype, []):
        raw = rec.get(col, "")
        if not raw:
            continue
        val = tex_verbatim(raw, where(col)) if kind == "verbatim" else tex_text(raw, where(col))
        if kind == "name":
            val = "{" + val + "}"
        fields.append((key, val))

    lines = [f"@{etype}{{{tag},"]
    lines += [_wrap_field(k, v) for k, v in fields]
    lines[-1] = lines[-1].rstrip(",")
    lines.append("}")
    return "\n".join(lines)

def _validate_bib_output(entries: list[str]):
    """Safety net: biber chokes on non-ASCII and on bare % / & in field values.
    Catches any conversion gap before LaTeX does, with the offending entry."""
    for entry in entries:
        m = re.match(r"@\w+\{([^,]+),", entry)
        key = m.group(1) if m else "?"
        for i, line in enumerate(entry.splitlines(), 1):
            bad = next((ch for ch in line if ord(ch) > 127), None)
            if bad is not None:
                sys.exit(f"BIB OUTPUT ERROR  entry '{key}' line {i}: "
                         f"non-ASCII {bad!r}: {line!r}")
            if re.search(r"(?<!\\)%", line):
                sys.exit(f"BIB OUTPUT ERROR  entry '{key}' line {i}: bare '%': {line!r}")
            if re.search(r"(?<!\\)&", line):
                sys.exit(f"BIB OUTPUT ERROR  entry '{key}' line {i}: bare '&': {line!r}")

def build_bib(wb) -> int:
    pubs = load_publications(wb)
    entries = [e for rec in pubs if (e := make_bibtex_entry(rec))]
    _validate_bib_output(entries)
    BIB.write_text("\n\n".join(entries) + "\n", encoding="utf-8")
    print(f"  Wrote {len(entries)} entries → {BIB.name}")
    return len(entries)

# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    today = date.today().strftime("%Y-%m-%d")
    print("Fetching spreadsheet from Google …")
    wb = sheet_data.workbook()

    tex = TEX.read_text(encoding="utf-8")

    print("\n── List sections (CV Data) ──────────────────────────────────────")
    tex = update_list_sections(tex, wb)

    print("\n── \\nocite{} blocks (Publications) ──────────────────────────────")
    tex = update_nocites(tex, wb, today)

    TEX.write_text(tex, encoding="utf-8")
    print(f"  Wrote {TEX.name}")

    print("\n── colonius.bib (Publications) ───────────────────────────────────")
    build_bib(wb)

    print("\nDone.")

if __name__ == "__main__":
    main()
